"""
Extract skill matrix data from Excel and convert to JSON for the web app.
Run this script whenever the Excel source data changes.
"""
import json
import openpyxl
import sys
import os

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'Downloads', 'Skill-Matrix-Software-Engineering_0.2 (1).xlsx')
# Also check same directory
if not os.path.exists(EXCEL_PATH):
    EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'Skill-Matrix-Software-Engineering_0.2 (1).xlsx')

OUTPUT_PATH = os.path.join(os.path.dirname(__file__), 'data.json')


def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()


def merge_employee_rows(raw_employees):
    """Merge multiple entries for the same employee (by name) into one."""
    merged = {}
    for emp in raw_employees:
        key = emp["name"]
        if key not in merged:
            merged[key] = emp
        else:
            existing = merged[key]
            if not existing["email"] and emp["email"]:
                existing["email"] = emp["email"]
            if not existing["product"] and emp["product"]:
                existing["product"] = emp["product"]
            for cat, skills in emp["skills"].items():
                if cat not in existing["skills"]:
                    existing["skills"][cat] = []
                existing_skill_names = {s["skill"] for s in existing["skills"][cat]}
                for s in skills:
                    if s["skill"] not in existing_skill_names:
                        existing["skills"][cat].append(s)
            for cert in emp.get("certifications", []):
                if cert not in existing["certifications"]:
                    existing["certifications"].append(cert)
    return list(merged.values())


def parse_dev_sheet(ws):
    """Parse the Dev sheet - multi-row per employee format."""
    employees = []
    current = None

    # Skill categories from header row 4
    categories = [
        {"name": "FrontEnd", "skillCol": 5, "profCol": 6},        # F, G
        {"name": ".Net", "skillCol": 7, "profCol": 8},             # H, I
        {"name": "Database", "skillCol": 9, "profCol": 10},        # J, K
        {"name": "Cloud", "skillCol": 11, "profCol": 12},          # L, M
        {"name": "Backend", "skillCol": 13, "profCol": 14},        # N, O
        {"name": "Deployment/Repository", "skillCol": 15, "profCol": 16},  # P, Q
        {"name": "Design Patterns", "skillCol": 17, "profCol": 18},       # R, S
        {"name": "Search Engines", "skillCol": 19, "profCol": 20},        # T, U
        {"name": "CRM (Integrations)", "skillCol": 21, "profCol": 22},    # V, W
        {"name": "UI Testing", "skillCol": 23, "profCol": 24},            # X, Y
    ]

    prev_name = ""
    for row_idx in range(6, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=c).value for c in range(1, 27)]

        product = safe_str(row[1])  # Column B
        name = safe_str(row[2])  # Column C
        email = safe_str(row[4])  # Column E
        certs = safe_str(row[25])  # Column Z

        # Skip fully empty rows
        if not name and not product:
            if current:
                employees.append(current)
                current = None
                prev_name = ""
            continue

        # Determine if this is a new employee: name changed from previous
        effective_name = name if name else prev_name
        is_new = (effective_name != prev_name) or current is None

        if is_new:
            if current:
                employees.append(current)
            current = {
                "name": effective_name,
                "product": product,
                "email": email if email else "",
                "department": "Dev",
                "skills": {},
                "certifications": []
            }
            for cat in categories:
                current["skills"][cat["name"]] = []

        prev_name = effective_name

        if current is None:
            continue

        # Update email if found in this row
        if email and not current["email"]:
            current["email"] = email

        # Update product if found
        if product and not current["product"]:
            current["product"] = product

        # Parse skills for each category
        for cat in categories:
            skill_raw = safe_str(row[cat["skillCol"]])
            prof = safe_str(row[cat["profCol"]])
            if skill_raw and skill_raw.strip() and prof and prof not in ("None/Low", ""):
                # Handle comma-separated skills in one cell
                skill_names = [s.strip() for s in skill_raw.split(",") if s.strip()]
                existing_skills = {s["skill"] for s in current["skills"][cat["name"]]}
                for skill in skill_names:
                    if skill not in existing_skills:
                        current["skills"][cat["name"]].append({
                            "skill": skill,
                            "proficiency": prof
                        })
                        existing_skills.add(skill)

        # Certifications
        if certs:
            for cert in certs.split(","):
                cert = cert.strip()
                if cert and cert not in current["certifications"]:
                    current["certifications"].append(cert)

    if current:
        employees.append(current)

    return merge_employee_rows(employees)


def parse_qa_sheet(ws):
    """Parse the QA sheet."""
    employees = []
    current = None

    categories = [
        {"name": "Automation", "skillCol": 4, "profCol": 5},      # E, F
        {"name": "Manual Testing", "skillCol": 6, "profCol": 7},  # G, H
        {"name": "API Testing", "skillCol": 8, "profCol": 9},     # I, J
        {"name": "Performance Testing", "skillCol": 10, "profCol": 11},  # K, L
        {"name": "CI/CD", "skillCol": 12, "profCol": 13},         # M, N
    ]

    for row_idx in range(6, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=c).value for c in range(1, 16)]

        sr_no = row[0]
        product = safe_str(row[1])
        name = safe_str(row[2])
        email = safe_str(row[3])
        certs = safe_str(row[14]) if len(row) > 14 else ""

        if not name and not product:
            continue

        is_new = False
        if sr_no is not None and str(sr_no).strip():
            is_new = True
        elif current and name and name != current["name"]:
            is_new = True

        if is_new:
            if current:
                employees.append(current)
            current = {
                "name": name,
                "product": product,
                "email": email if email else "",
                "department": "QA",
                "skills": {},
                "certifications": []
            }
            for cat in categories:
                current["skills"][cat["name"]] = []

        if current is None:
            continue

        if email and not current["email"]:
            current["email"] = email
        if product and not current["product"]:
            current["product"] = product

        for cat in categories:
            skill = safe_str(row[cat["skillCol"]])
            prof = safe_str(row[cat["profCol"]])
            if skill and skill.strip() and prof and prof not in ("None/Low", ""):
                current["skills"][cat["name"]].append({
                    "skill": skill,
                    "proficiency": prof
                })

        if certs:
            for cert in certs.split(","):
                cert = cert.strip()
                if cert and cert not in current["certifications"]:
                    current["certifications"].append(cert)

    if current:
        employees.append(current)

    return merge_employee_rows(employees)


def parse_cloud_sheet(ws):
    """Parse Cloud Ops sheet."""
    employees = []
    current = None

    categories = [
        {"name": "Cloud", "skillCol": 5, "profCol": 6},                  # F, G
        {"name": "Deployment/Repository", "skillCol": 7, "profCol": 8},  # H, I
        {"name": "DevOps Tool", "skillCol": 9, "profCol": 10},           # J, K
        {"name": "Administrator (OS)/Storage", "skillCol": 11, "profCol": 12},  # L, M
    ]

    for row_idx in range(6, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=c).value for c in range(1, 15)]

        sr_no = row[0]
        product = safe_str(row[1])
        name = safe_str(row[2])
        email = safe_str(row[4])
        certs = safe_str(row[13]) if len(row) > 13 else ""

        if not name and not product and not any(safe_str(row[c]) for c in range(5, 13)):
            continue

        is_new = False
        if sr_no is not None and str(sr_no).strip():
            is_new = True
        elif current and name and name != current["name"]:
            is_new = True

        if is_new:
            if current:
                employees.append(current)
            current = {
                "name": name if name else "",
                "product": product,
                "email": email if email else "",
                "department": "Cloud Ops",
                "skills": {},
                "certifications": []
            }
            for cat in categories:
                current["skills"][cat["name"]] = []

        if current is None:
            continue

        if email and not current["email"]:
            current["email"] = email
        if product and not current["product"]:
            current["product"] = product

        for cat in categories:
            skill = safe_str(row[cat["skillCol"]])
            prof = safe_str(row[cat["profCol"]])
            if skill and skill.strip() and prof and prof not in ("None/Low", ""):
                current["skills"][cat["name"]].append({
                    "skill": skill,
                    "proficiency": prof
                })

        if certs:
            for cert in certs.split(","):
                cert = cert.strip()
                if cert and cert not in current["certifications"]:
                    current["certifications"].append(cert)

    if current:
        employees.append(current)

    return merge_employee_rows(employees)


def parse_simple_sheet(ws, department, name_col=3, email_col=4, product_col=2,
                       skill_col=5, prof_col=6, cert_col=7, start_row=2,
                       header_rows=None, extra_skill_cols=None):
    """Generic parser for simple row-per-skill sheets like Design, Project Management, etc.
    Each row: product, name, email, skill, proficiency, certs.
    Name/email/product may only appear on the first row (merged cells) — continuation
    rows carry skills for the same employee.
    extra_skill_cols: list of (skill_col, prof_col) tuples for additional skill columns (e.g. Technical Writing).
    """
    employees = []
    current = None
    if header_rows is None:
        header_rows = set()

    for row_idx in range(start_row, ws.max_row + 1):
        if row_idx in header_rows:
            continue
        max_col = max(cert_col + 1, (max(c for pair in (extra_skill_cols or []) for c in pair) + 1) if extra_skill_cols else cert_col + 1)
        row = [ws.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)]

        name = safe_str(row[name_col - 1])
        email = safe_str(row[email_col - 1])
        product = safe_str(row[product_col - 1])
        skill = safe_str(row[skill_col - 1])
        prof = safe_str(row[prof_col - 1])
        cert = safe_str(row[cert_col - 1]) if cert_col <= len(row) else ""

        # If there's a name, start or find this employee
        if name:
            # Skip rows where name looks like an email (data entry error)
            if "@" in name:
                continue
            # Find existing or create new employee
            found = None
            for emp in employees:
                if emp["name"] == name:
                    found = emp
                    break
            if found is None:
                found = {
                    "name": name,
                    "product": product,
                    "email": email,
                    "department": department,
                    "skills": {department: []},
                    "certifications": []
                }
                employees.append(found)
            if email and not found["email"]:
                found["email"] = email
            if product and not found["product"]:
                found["product"] = product
            current = found
        else:
            # Continuation row (no name) — belongs to the current employee
            if current is None:
                continue

        # Skip rows where skill looks like an email (data entry error)
        if skill and "@" in skill:
            continue

        # If no skill and no name, skip entirely
        if not skill and not name:
            continue

        # Collect all skills from this row (main + extra columns)
        row_skills = []
        if skill and prof:
            # Handle comma-separated skills
            for s in skill.split(","):
                s = s.strip()
                if s:
                    row_skills.append({"skill": s, "proficiency": prof})

        if extra_skill_cols:
            for sc, pc in extra_skill_cols:
                es = safe_str(row[sc - 1]) if sc <= len(row) else ""
                ep = safe_str(row[pc - 1]) if pc <= len(row) else ""
                if es and ep:
                    for s in es.split(","):
                        s = s.strip()
                        if s:
                            row_skills.append({"skill": s, "proficiency": ep})

        # Add skills to current employee
        existing_skill_names = {s["skill"] for s in current["skills"].get(department, [])}
        for sk in row_skills:
            if sk["skill"] not in existing_skill_names:
                current["skills"].setdefault(department, []).append(sk)
                existing_skill_names.add(sk["skill"])

        # Certs
        if cert:
            for c in cert.split(","):
                c = c.strip()
                if c and c not in current["certifications"]:
                    current["certifications"].append(c)

    return employees


def parse_reference_dev(ws):
    """Parse Reference (Dev) sheet for dropdown options."""
    categories = {}
    headers = {
        "Front-end": 4,       # D
        ".NET/Authentication": 6,  # F
        "Database": 8,         # H
        "Cloud": 10,           # J
        "Back End": 12,        # L
        "Deployment/Repository": 14,  # N
        "Design Patterns/Architecture": 16,  # P
        "Search Engines": 18,  # R
        "CRM (Integrations)": 20,  # T
        "Telephony Services": 22,  # V
        "Dev Ops Tool": 24,    # X
        "Administrator (OS)/Storage": 26,  # Z
    }

    for name, col in headers.items():
        skills = []
        for row_idx in range(4, 21):
            val = safe_str(ws.cell(row=row_idx, column=col).value)
            if val:
                skills.append(val)
        if skills:
            categories[name] = skills

    return categories


def parse_reference_qa(ws):
    """Parse Reference (QA) sheet for dropdown options."""
    categories = {}
    headers = {
        "Automation": 4,        # D
        "Manual Testing": 6,    # F
        "API Testing": 8,       # H
        "Performance Testing": 10,  # J
        "UI Testing": 12,       # L
    }

    for name, col in headers.items():
        skills = []
        for row_idx in range(4, 14):
            val = safe_str(ws.cell(row=row_idx, column=col).value)
            if val:
                skills.append(val)
        if skills:
            categories[name] = skills

    return categories


def main():
    path = EXCEL_PATH
    if len(sys.argv) > 1:
        path = sys.argv[1]

    print(f"Loading workbook from: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)

    all_departments = [
        "Dev", "QA", "Cloud Ops",
        "Design", "Project Management", "Product Management",
        "Technical Writing", "Product Marketing", "EIT"
    ]

    data = {
        "employees": [],
        "reference": {
            "dev": {},
            "qa": {},
            "proficiencyLevels": [
                {"level": 1, "name": "None/Low", "description": "Unable to perform; little to no experience"},
                {"level": 2, "name": "Basic", "description": "Limited in ability or knowledge; Cannot perform for critical tasks; Need significant help from others"},
                {"level": 3, "name": "Intermediate", "description": "Able to perform at basic level; Has some direct experience; Needs help from time to time"},
                {"level": 4, "name": "Expert", "description": "Capable and experienced; Able to work independently and need no assistance; Can lead and train others"}
            ]
        },
        "products": [],
        "departments": all_departments
    }

    # Parse employee data - original 3 teams
    dev_employees = parse_dev_sheet(wb["Dev"])
    qa_employees = parse_qa_sheet(wb["QA"])
    cloud_employees = parse_cloud_sheet(wb["Cloud Ops"])

    # Parse new sheets
    # Design: header row 1, data from row 2, cols: A=Sr.No, B=Product, C=Name, D=Email, E=Skills, F=Proficiency, G=Certs
    design_employees = parse_simple_sheet(
        wb["Design"], "Design",
        name_col=3, email_col=4, product_col=2,
        skill_col=5, prof_col=6, cert_col=7, start_row=2
    )

    # Project Management: header row 2, data from row 3
    pm_employees = parse_simple_sheet(
        wb["Project Management"], "Project Management",
        name_col=3, email_col=4, product_col=2,
        skill_col=5, prof_col=6, cert_col=7, start_row=3
    )

    # Product Management: header row 3, data from row 4
    prodmgmt_employees = parse_simple_sheet(
        wb["Product Management"], "Product Management",
        name_col=3, email_col=4, product_col=2,
        skill_col=5, prof_col=6, cert_col=7, start_row=4
    )

    # Technical Writing: header rows 1-2, data from row 3
    # TWO skill column pairs: E/F (TW Tools) and G/H (Other Tools), Certs in I
    tw_employees = parse_simple_sheet(
        wb["Technical Writing"], "Technical Writing",
        name_col=3, email_col=4, product_col=2,
        skill_col=5, prof_col=6, cert_col=9, start_row=3,
        extra_skill_cols=[(7, 8)]
    )

    # Product Marketing: header row 3, data from row 4
    prodmkt_employees = parse_simple_sheet(
        wb["Product Marketing"], "Product Marketing",
        name_col=3, email_col=4, product_col=2,
        skill_col=5, prof_col=6, cert_col=7, start_row=4
    )

    # EIT: NO product/sr.no columns. A=Name, B=Email, C=Skills, D=Proficiency. Data from row 2.
    eit_employees = parse_simple_sheet(
        wb["EIT"], "EIT",
        name_col=1, email_col=2, product_col=1,  # product_col=1 (name) as placeholder, no product col
        skill_col=3, prof_col=4, cert_col=99, start_row=2  # cert_col=99 to skip (no cert column)
    )
    # EIT has no product field, clear it
    for emp in eit_employees:
        emp["product"] = ""

    data["employees"] = (
        dev_employees + qa_employees + cloud_employees +
        design_employees + pm_employees + prodmgmt_employees +
        tw_employees + prodmkt_employees + eit_employees
    )

    # Parse reference data
    data["reference"]["dev"] = parse_reference_dev(wb["Reference (Dev)"])
    data["reference"]["qa"] = parse_reference_qa(wb["Reference (QA) "])

    # Extract unique products
    products = set()
    for emp in data["employees"]:
        if emp.get("product"):
            products.add(emp["product"])
    data["products"] = sorted(list(products))

    # Filter out employees with empty names
    data["employees"] = [e for e in data["employees"] if e.get("name")]

    print(f"Extracted {len(data['employees'])} employees")
    for dept in all_departments:
        count = len([e for e in data["employees"] if e["department"] == dept])
        print(f"  {dept}: {count}")
    print(f"Products: {data['products']}")

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"Data saved to: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
