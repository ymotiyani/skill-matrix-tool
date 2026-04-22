import openpyxl

wb = openpyxl.load_workbook('source.xlsx', data_only=True)

def safe_str(v):
    if v is None:
        return ""
    return str(v).strip()

# Debug Design sheet
ws = wb["Design"]
print(f"Design sheet: max_row={ws.max_row}, max_col={ws.max_column}")
for row_idx in range(1, ws.max_row + 1):
    row = [ws.cell(row=row_idx, column=c).value for c in range(1, 8)]
    name = safe_str(row[2])  # col C (index 2)
    skill = safe_str(row[4])  # col E (index 4)
    prof = safe_str(row[5])   # col F (index 5)
    print(f"  Row {row_idx}: name='{name}' skill='{skill}' prof='{prof}' | raw={row}")

print()

# Debug Project Management sheet
ws2 = wb["Project Management"]
print(f"Project Management sheet: max_row={ws2.max_row}, max_col={ws2.max_column}")
for row_idx in range(1, min(20, ws2.max_row + 1)):
    row = [ws2.cell(row=row_idx, column=c).value for c in range(1, 9)]
    name = safe_str(row[2])
    skill = safe_str(row[4])
    prof = safe_str(row[5])
    print(f"  Row {row_idx}: name='{name}' skill='{skill}' prof='{prof}' | raw={row}")

print()

# Debug EIT sheet
ws3 = wb["EIT"]
print(f"EIT sheet: max_row={ws3.max_row}, max_col={ws3.max_column}")
for row_idx in range(1, ws3.max_row + 1):
    row = [ws3.cell(row=row_idx, column=c).value for c in range(1, 5)]
    name = safe_str(row[0])  # col A
    skill = safe_str(row[2])  # col C
    prof = safe_str(row[3])   # col D
    print(f"  Row {row_idx}: name='{name}' skill='{skill}' prof='{prof}' | raw={row}")
